"""Импорт и экспорт задач в Excel.

Поддерживаемые форматы импорта:
- .xlsx (openpyxl)
- старый бинарный .xls (xlrd)
- «.xls» из Jira, который на самом деле HTML-таблица

Строка заголовков ищется автоматически (первая строка, где есть Summary /
Название). Неизвестные столбцы пропускаются.
"""
import io
import re
from datetime import datetime
from html.parser import HTMLParser

import openpyxl
import xlrd
from markupsafe import escape

from .extensions import db
from .history import add_event
from .models import (ISSUE_TYPES, PARENT_TYPE, Component, Customer, Issue,
                     Project, Sprint, Status, Team, User)

# Заголовок столбца (lower) -> наше поле
COLUMN_MAP = {
    'project': 'project',
    'проект': 'project',
    'key': 'key',
    'summary': 'title',
    'название': 'title',
    'issue type': 'type',
    'тип': 'type',
    'status': 'status',
    'статус': 'status',
    'assignee': 'assignee',
    'исполнитель': 'assignee',
    'reporter': 'reporter',
    'автор': 'reporter',
    'team': 'team',
    'команда': 'team',
    'customer': 'customer',
    'заказчик': 'customer',
    'component': 'component',
    'components': 'component',
    'компонент': 'component',
    'sprint': 'sprint',
    'спринт': 'sprint',
    'created': 'created',
    'создана': 'created',
    'updated': 'updated',
    'обновлена': 'updated',
    'description': 'description',
    'описание': 'description',
    'epic link': 'epic_link',
    'epic name': 'epic_name',
    'priority': 'priority',
    'приоритет': 'priority',
}

TYPE_MAP = {
    'epic': 'epic',
    'feature': 'feature',
    'story': 'feature',
    'task': 'task',
    'sub-task': 'task',
    'подзадача': 'task',
    'improvement': 'task',
    'bug': 'bug',
    'баг': 'bug',
}

# Приоритеты Jira (обе стандартные схемы) -> наши
PRIORITY_MAP = {
    'blocker': 'critical',
    'critical': 'critical',
    'highest': 'highest',
    'high': 'high',
    'major': 'high',
    'medium': 'normal',
    'normal': 'normal',
    'minor': 'low',
    'low': 'low',
    'lowest': 'low',
    'trivial': 'low',
}

DONE_STATUS_WORDS = {'done', 'closed', 'resolved', 'готово', 'закрыто', 'закрыт'}


def _looks_done(status_name):
    """Считает статус закрывающим, если в названии есть «done»-слово.

    Ловит и составные названия вроде «Bug. Done» или «Закрыто (дубль)».
    """
    words = set(re.findall(r'[a-zа-яё]+', status_name.lower()))
    return bool(words & DONE_STATUS_WORDS)

DATE_FORMATS = ('%d/%b/%y %I:%M %p', '%d/%b/%Y %I:%M %p',
                '%d.%m.%Y %H:%M', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d')

# Мусорные значения Jira, которые не считаем данными
JUNK_VALUES = {'unassigned', 'unresolved', 'no permission'}
LEXORANK_RE = re.compile(r'^\d+\|')  # значения вида «0|i01c7r:» (ранги Jira)


# ---------- Чтение файлов ----------

class _HTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self._row = None
        self._cell = None

    def handle_starttag(self, tag, attrs):
        if tag == 'tr':
            self._row = []
        elif tag in ('td', 'th') and self._row is not None:
            self._cell = []

    def handle_endtag(self, tag):
        if tag == 'tr' and self._row is not None:
            self.rows.append(self._row)
            self._row = None
        elif tag in ('td', 'th') and self._cell is not None:
            self._row.append(re.sub(r'\s+', ' ', ''.join(self._cell)).strip())
            self._cell = None

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)


def read_rows(data, filename=''):
    """Возвращает таблицу (список списков) из xlsx / xls / jira-html файла."""
    head = data[:8].lstrip()
    if data[:4] == b'\xd0\xcf\x11\xe0':  # OLE2 -> бинарный .xls
        book = xlrd.open_workbook(file_contents=data)
        sheet = book.sheet_by_index(0)
        rows = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                elif cell.ctype == xlrd.XL_CELL_NUMBER and cell.value == int(cell.value):
                    row.append(int(cell.value))
                else:
                    row.append(cell.value)
            rows.append(row)
        return rows
    if head[:1] == b'<':  # HTML-таблица из Jira
        parser = _HTMLTableParser()
        parser.feed(data.decode('utf-8', errors='replace'))
        return parser.rows
    # иначе считаем, что это xlsx (zip)
    book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = book.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def find_header(rows):
    """Ищет строку заголовков. Возвращает (номер строки, {поле: индекс})."""
    for idx, row in enumerate(rows[:20]):
        lowered = [str(c).strip().lower() if c is not None else '' for c in row]
        if 'summary' in lowered or 'название' in lowered:
            colmap = {}
            for col_idx, header in enumerate(lowered):
                field = COLUMN_MAP.get(header)
                if field and field not in colmap:
                    colmap[field] = col_idx
            return idx, colmap
    raise ValueError('Не нашла строку заголовков: нужен столбец Summary или Название.')


# ---------- Импорт ----------

def _cell(row, colmap, field):
    idx = colmap.get(field)
    if idx is None or idx >= len(row):
        return None
    value = row[idx]
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text or text.lower() in JUNK_VALUES:
        return None
    return text


def _parse_date(value):
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def import_rows(rows, current_user, dry_run=False):
    """Создаёт задачи из таблицы. Возвращает статистику и предупреждения.

    При dry_run=True ничего не сохраняет (транзакция откатывается), но
    возвращает ту же статистику — для предпросмотра перед импортом.
    """
    header_idx, colmap = find_header(rows)
    if 'title' not in colmap:
        raise ValueError('Не найден столбец Summary / Название.')

    warnings = set()
    created = 0
    skipped = 0

    # Кеши справочников (по lower-имени)
    def _lookup_cache(model):
        return {obj.name.strip().lower(): obj for obj in model.query.all()}

    caches = {
        'project': (_lookup_cache(Project), Project),
        'team': (_lookup_cache(Team), Team),
        'customer': (_lookup_cache(Customer), Customer),
        'component': (_lookup_cache(Component), Component),
    }
    statuses = _lookup_cache(Status)
    sprints = _lookup_cache(Sprint)
    users = {u.name.strip().lower(): u for u in User.query.all()}
    # Эпики для привязки по Epic Link: уже существующие + созданные этим импортом
    epics = {i.title.strip().lower(): i
             for i in Issue.query.filter_by(type='epic').all()}
    pending_epic_links = []  # (issue, значение Epic Link)
    next_position = (db.session.query(db.func.max(Status.position)).scalar() or 0) + 1

    def get_or_create_dict(kind, name):
        cache, model = caches[kind]
        obj = cache.get(name.lower())
        if not obj:
            obj = model(name=name)
            db.session.add(obj)
            db.session.flush()
            cache[name.lower()] = obj
        return obj

    for row in rows[header_idx + 1:]:
        title = _cell(row, colmap, 'title')
        if not title:
            skipped += 1
            continue

        issue = Issue(title=str(title)[:300])

        raw_type = (_cell(row, colmap, 'type') or 'task')
        issue.type = TYPE_MAP.get(str(raw_type).lower(), 'task')
        if str(raw_type).lower() not in TYPE_MAP:
            warnings.add(f'Неизвестный тип «{raw_type}» — импортирован как Task.')

        raw_priority = _cell(row, colmap, 'priority')
        if raw_priority:
            issue.priority = PRIORITY_MAP.get(str(raw_priority).lower(), 'normal')
            if str(raw_priority).lower() not in PRIORITY_MAP:
                warnings.add(f'Неизвестный приоритет «{raw_priority}» — записан Normal.')

        # Статус: создаём при необходимости
        status_name = _cell(row, colmap, 'status')
        status_obj = None
        if status_name:
            key = str(status_name).lower()
            status_obj = statuses.get(key)
            if not status_obj:
                status_obj = Status(
                    name=str(status_name)[:80], position=next_position,
                    is_done=_looks_done(str(status_name)))
                next_position += 1
                db.session.add(status_obj)
                db.session.flush()
                statuses[key] = status_obj
        if not status_obj:
            status_obj = (Status.query.order_by(Status.position).first())
            if not status_obj:
                raise ValueError('В системе нет ни одного статуса.')
        issue.status_id = status_obj.id

        # Справочники
        for kind, attr in (('project', 'project_id'), ('team', 'team_id'),
                           ('customer', 'customer_id'),
                           ('component', 'component_id')):
            name = _cell(row, colmap, kind)
            if name:
                setattr(issue, attr, get_or_create_dict(kind, str(name)[:120]).id)

        # Пользователи: ищем по имени, не создаём
        reporter_name = _cell(row, colmap, 'reporter')
        reporter = users.get(str(reporter_name).lower()) if reporter_name else None
        if reporter_name and not reporter:
            warnings.add(f'Автор «{reporter_name}» не найден — автором записан текущий пользователь.')
        issue.reporter_id = (reporter or current_user).id

        assignee_name = _cell(row, colmap, 'assignee')
        if assignee_name:
            assignee = users.get(str(assignee_name).lower())
            if assignee:
                issue.assignee_id = assignee.id
            else:
                warnings.add(f'Исполнитель «{assignee_name}» не найден — оставлен пустым.')

        # Спринт (отбрасываем джировские ранги вида «0|i01c7r:»)
        sprint_name = _cell(row, colmap, 'sprint')
        if sprint_name and not LEXORANK_RE.match(str(sprint_name)):
            key = str(sprint_name).lower()
            sprint = sprints.get(key)
            if not sprint:
                sprint = Sprint(name=str(sprint_name)[:120])
                db.session.add(sprint)
                db.session.flush()
                sprints[key] = sprint
            issue.sprint_id = sprint.id

        # Описание: плоский текст -> простой HTML
        description = _cell(row, colmap, 'description')
        if description:
            paragraphs = str(description).splitlines()
            issue.summary = ''.join(
                f'<p>{escape(p)}</p>' for p in paragraphs if p.strip())

        created_at = _parse_date(_cell(row, colmap, 'created'))
        updated_at = _parse_date(_cell(row, colmap, 'updated'))
        if created_at:
            issue.created_at = created_at
        issue.updated_at = updated_at or created_at or datetime.utcnow()

        db.session.add(issue)
        db.session.flush()
        add_event(issue, current_user, 'created')
        created += 1

        if issue.type == 'epic':
            epics.setdefault(issue.title.strip().lower(), issue)
            # В Jira Epic Link ссылается на Epic Name, который может
            # отличаться от Summary — запоминаем оба варианта.
            epic_name = _cell(row, colmap, 'epic_name')
            if epic_name:
                epics.setdefault(str(epic_name).strip().lower(), issue)

        epic_link = _cell(row, colmap, 'epic_link')
        if epic_link:
            pending_epic_links.append((issue, str(epic_link).strip()))

    # Привязка к эпикам: после основного цикла, потому что эпик может
    # встретиться в файле позже своих задач.
    epic_linked = 0
    epic_link_rejected = []
    for issue, epic_name in pending_epic_links:
        epic = epics.get(epic_name.lower())
        if not epic:
            warnings.add(f'Эпик «{epic_name}» не найден — привязка пропущена.')
            continue
        if PARENT_TYPE.get(issue.type) == 'epic':
            issue.parent_id = epic.id
            epic_linked += 1
        else:
            epic_link_rejected.append({
                'id': issue.id,
                'title': issue.title,
                'type': ISSUE_TYPES.get(issue.type, issue.type),
                'epic': epic.title,
            })

    if dry_run:
        db.session.rollback()
    else:
        db.session.commit()
    return {'created': created, 'skipped': skipped, 'warnings': sorted(warnings),
            'epic_linked': epic_linked, 'epic_link_rejected': epic_link_rejected}


# ---------- Экспорт ----------

EXPORT_HEADERS = ['Key', 'Project', 'Summary', 'Issue Type', 'Priority',
                  'Status', 'Assignee', 'Reporter', 'Team', 'Customer',
                  'Component', 'Sprint', 'Parent', 'Created', 'Updated',
                  'Description']


def _plain_text(html):
    if not html:
        return ''
    text = re.sub(r'</p>\s*<p>', '\n', html)
    text = re.sub(r'<br\s*/?>', '\n', text)
    return re.sub(r'<[^>]+>', '', text).strip()


def build_export(issues):
    """Собирает xlsx c задачами, возвращает BytesIO."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'Issues'
    sheet.append(EXPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for issue in issues:
        sheet.append([
            issue.id,
            issue.project.name if issue.project else '',
            issue.title,
            issue.type_label,
            issue.priority_label,
            issue.status.name,
            issue.assignee.name if issue.assignee else '',
            issue.reporter.name,
            issue.team.name if issue.team else '',
            issue.customer.name if issue.customer else '',
            issue.component.name if issue.component else '',
            issue.sprint.name if issue.sprint else '',
            f'#{issue.parent_id}' if issue.parent_id else '',
            issue.created_at.strftime('%d.%m.%Y %H:%M'),
            issue.updated_at.strftime('%d.%m.%Y %H:%M'),
            _plain_text(issue.summary),
        ])

    widths = [8, 18, 50, 10, 10, 14, 20, 20, 15, 15, 15, 15, 8, 17, 17, 60]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    output = io.BytesIO()
    book.save(output)
    output.seek(0)
    return output
